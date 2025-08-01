# views.py
from django.shortcuts import render
from django.http import HttpResponse
from .forms import DateRangeForm
from cupeinter.models import CopeInter
import openpyxl

def export_filtered_excel(request):
    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start = form.cleaned_data['start_date']
            end = form.cleaned_data['end_date']

            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

            def add_sheet(name, queryset, fields):
                sheet = workbook.create_sheet(title=name)
                for i, field in enumerate(fields, 1):
                    sheet.cell(row=1, column=i, value=field)
                for row_idx, obj in enumerate(queryset, 2):
                    for col_idx, field in enumerate(fields, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=str(getattr(obj, field)))

            # فیلتر بر اساس تاریخ
            data_qs = CopeInter.objects.filter(created_at__range=[start, end])
            # customer_qs = Customer.objects.filter(created__range=[start, end])
            # order_qs = Order.objects.filter(created__range=[start, end])

            add_sheet('Data', data_qs, ['id', 'name', 'age', 'email', 'created_at'])
            # add_sheet('Customer', customer_qs, ['id', 'full_name', 'email', 'created'])
            # add_sheet('Order', order_qs, ['id', 'customer_id', 'total', 'created'])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=filtered_export.xlsx'
            workbook.save(response)
            return response
    else:
        form = DateRangeForm()

    return render(request, 'export_form.html', {'form': form})
